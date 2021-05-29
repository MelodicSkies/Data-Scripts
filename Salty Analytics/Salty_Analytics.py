from openpyxl import load_workbook
from datetime import datetime


sales_data = load_workbook(filename = "Salty Analytics Intership Sample Dataset.xlsx", data_only = True)

daily_sales = sales_data["total sales"]
total_sales = 1
sales_by_date = []

for i in range(2, daily_sales.max_row, 1):

    if i != 2:
        if daily_sales.cell(row = i, column = 6).value.date() == daily_sales.cell(row = i - 1, column = 6).value.date():
            total_sales += 1
            try: 
                if daily_sales.cell(row = i, column = 6).value.date() != daily_sales.cell(row = i + 1, column = 6).value.date():
                    sales_by_date.append((daily_sales.cell(row = i, column = 6).value.date(), total_sales))
                    
            except:
                sales_by_date.append((daily_sales.cell(row = i, column = 6).value.date(), total_sales))
              
        else:
            total_sales = 1

for i in range(0, len(sales_by_date),1):

    tuple = sales_by_date[i]
    daily_sales.cell(row = i+2, column = 10).value = tuple[0]
    daily_sales.cell(row = i+2, column = 11).value = tuple[1]
 
sales_data.save("Salty Analytics Intership Sample Dataset.xlsx")



    